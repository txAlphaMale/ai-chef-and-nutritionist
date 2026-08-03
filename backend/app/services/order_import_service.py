"""Backlog B10.3 (author-requested group, 2026-08-01): generic order-
history CSV/XLSX importer -- the recommended build from this session's
Walmart-integration research, which found no public consumer purchase-
history API and no stable, documented export schema (every real-world
export comes from an unofficial third-party browser extension, each
with its own, mutable column layout). Rather than guess a "Walmart"
column set against an unverified schema, this module builds a fully
generic tabular importer: parse whatever CSV/XLSX the user actually has,
suggest a column mapping via keyword matching, let the user confirm or
correct it, and optionally save it as a named profile for next time.

Deliberately pure/parsing-only, no AI, no network call -- unlike the
receipt/list import (B4.2), a spreadsheet with known column headers
needs no model to interpret it. Shares the review-then-confirm intake
discipline anyway: this module only produces a PREVIEW
(`VisionDetectedItem`-shaped, same as every other intake source in this
app), never writes to inventory directly. See routers/inventory.py's
order-import endpoints for where this is wired in, and
app/models/inventory.py's `OrderImportProfile` docstring for why no
pre-built retailer profile ships.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime

import openpyxl

from app.schemas.inventory import ColumnMapping, VisionDetectedItem
from app.services import package_parsing

# Checked in priority order per field -- the FIRST header (from the ones
# not already claimed by an earlier field) whose text exactly matches or
# contains one of these substrings (case-insensitive) is suggested.
# Fields are resolved in this fixed order specifically to resolve
# collisions between two field's keyword lists: name_column is resolved
# first against a specific "item/product name" style phrase (not a bare
# "item" substring) so it doesn't grab a header meant for another field;
# price_column is resolved before unit_column so a "Unit Price" header
# (which contains "unit" but means price-per-unit) lands on price_column
# rather than unit_column -- caught by a test written against exactly
# this header text, not assumed.
_KEYWORDS: dict[str, list[str]] = {
    "name_column": ["item name", "product name", "description", "product", "item"],
    "quantity_column": ["quantity", "qty", "count"],
    "price_column": ["price", "cost", "amount", "total", "subtotal", "paid"],
    "unit_column": ["unit", "uom", "measure"],
    "date_column": ["order date", "purchase date", "date"],
}

# Tried in order; the first that parses the whole cell wins. Deliberately
# a short, common list rather than an exhaustive one -- an unparseable
# date is left as None (never guessed) rather than silently misread.
_DATE_FORMATS = ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%B %d, %Y", "%b %d, %Y"]

_CURRENCY_STRIP = re.compile(r"[^0-9.\-]")


def guess_column_mapping(headers: list[str]) -> ColumnMapping:
    """Best-effort suggestion only -- always returned alongside the
    headers so the user can see and correct it before anything is
    parsed for real (see OrderImportPreviewResponse)."""
    available = list(headers)
    result: dict[str, str | None] = {}
    for field, keywords in _KEYWORDS.items():
        chosen = None
        for keyword in keywords:
            for header in available:
                if keyword in header.lower():
                    chosen = header
                    break
            if chosen:
                break
        result[field] = chosen
        if chosen:
            available.remove(chosen)
    return ColumnMapping(**result)


def parse_tabular_file(raw_bytes: bytes, filename: str, content_type: str) -> tuple[list[str], list[dict[str, str]]]:
    """Returns (headers, rows) -- rows are dicts keyed by header text
    (not position), since column order isn't guaranteed stable across
    re-exports of the same source but header text usually is."""
    lower_name = (filename or "").lower()
    is_csv = content_type in ("text/csv", "application/csv") or lower_name.endswith(".csv")
    is_xlsx = (
        content_type
        in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel")
        or lower_name.endswith(".xlsx")
        or lower_name.endswith(".xls")
    )

    if is_csv:
        # utf-8-sig quietly strips a BOM if present (common in
        # Excel-produced CSV exports) rather than leaving it stuck to
        # the first header's name.
        text = raw_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = [h.strip() for h in (reader.fieldnames or [])]
        rows = [{(k or "").strip(): (v or "").strip() for k, v in row.items()} for row in reader]
        return headers, rows

    if is_xlsx:
        workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        # Prefer a sheet literally named "Items" if present -- the
        # Walmart Invoice Exporter extension's modern export format uses
        # a two-sheet "Orders"+"Items" workbook, and "Items" is the
        # sheet with one row per purchased line, which is what this
        # importer wants. Falls back to the first sheet for any other
        # single-sheet export.
        sheet = None
        for name in workbook.sheetnames:
            if name.strip().lower() == "items":
                sheet = workbook[name]
                break
        if sheet is None:
            sheet = workbook[workbook.sheetnames[0]]

        row_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            return [], []
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        rows = []
        for raw_row in row_iter:
            row = {}
            # strict=False, deliberately: a spreadsheet row that is short
            # or long relative to the header row is normal in real
            # exports, and the surrounding code already skips blank
            # headers. Truncating to the shorter of the two is the
            # intended behaviour, not an oversight.
            for header, value in zip(headers, raw_row, strict=False):
                if not header:
                    continue
                row[header] = "" if value is None else str(value).strip()
            if any(v for v in row.values()):  # skip fully blank rows
                rows.append(row)
        return headers, rows

    raise ValueError("Unsupported file type for order-history import -- upload a .csv or .xlsx file.")


def _parse_quantity(raw: str | None) -> tuple[float, str | None]:
    if not raw:
        return 1.0, None
    try:
        return float(_CURRENCY_STRIP.sub("", raw)), None
    except ValueError:
        return 1.0, f"could not parse quantity value {raw!r}, defaulted to 1"


def _parse_price(raw: str | None) -> tuple[float | None, str | None]:
    if not raw:
        return None, None
    cleaned = _CURRENCY_STRIP.sub("", raw)
    try:
        return float(cleaned), None
    except ValueError:
        return None, f"could not parse price value {raw!r}"


def _parse_date(raw: str | None) -> date | None:
    if not raw:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def apply_mapping(
    headers: list[str], rows: list[dict[str, str]], mapping: ColumnMapping
) -> tuple[list[VisionDetectedItem], int]:
    """Turns raw parsed rows into the shared preview shape using the
    given mapping. Returns (items, skipped_row_count) -- a row with no
    usable name (blank spacer row, a subtotal/tax/footer line with
    nothing under the mapped name column) is skipped rather than
    creating a nameless inventory item, same discipline as the receipt/
    list AI import's own "skip non-item lines" instruction."""
    items: list[VisionDetectedItem] = []
    skipped = 0
    for row in rows:
        name = (row.get(mapping.name_column, "") if mapping.name_column else "").strip()
        if not name:
            skipped += 1
            continue

        raw_qty = row.get(mapping.quantity_column) if mapping.quantity_column else None
        quantity, qty_note = _parse_quantity(raw_qty)
        if mapping.quantity_column is None:
            qty_note = "no quantity column mapped, defaulted to 1"

        raw_unit = (row.get(mapping.unit_column) or None) if mapping.unit_column else None

        # Package/measurement split (2026-08-02) -- a spreadsheet's own
        # "unit" column is exactly as likely to contain a compound size
        # string ("8 oz", "500g") as a receipt line is, so it gets the
        # same best-effort split. `quantity` from the mapped quantity
        # column is treated as "how many purchased" (package_count);
        # when the unit column also yields a package size, the final
        # on-hand quantity is their product, same convention
        # inventory_service.parse_vision_response uses for the AI import
        # paths -- see package_parsing.py's module docstring.
        unit = raw_unit
        package_quantity = None
        package_count = quantity
        package_descriptor = None
        final_quantity = quantity
        parsed_package = package_parsing.parse_package_text(raw_unit)
        if parsed_package is not None:
            unit = parsed_package.unit
            package_quantity = parsed_package.package_quantity
            package_descriptor = parsed_package.package_descriptor
            package_count = quantity * parsed_package.package_count
            final_quantity = package_count * package_quantity

        raw_price = row.get(mapping.price_column) if mapping.price_column else None
        price, price_note = _parse_price(raw_price)

        raw_date = row.get(mapping.date_column) if mapping.date_column else None
        purchased = _parse_date(raw_date)

        notes = [n for n in (qty_note, price_note) if n]
        items.append(
            VisionDetectedItem(
                name=name,
                estimated_quantity=final_quantity,
                unit=unit,
                package_quantity=package_quantity,
                package_count=package_count,
                package_descriptor=package_descriptor,
                category="other",
                confidence_note="; ".join(notes) or None,
                unit_price=price,
                purchased_date=purchased,
            )
        )
    return items, skipped
