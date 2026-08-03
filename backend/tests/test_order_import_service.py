"""Tests for the B10.3 generic order-history CSV/XLSX importer
(order_import_service.py). No AI/network dependency here at all -- this
module is pure deterministic parsing, so unlike B4.2's receipt import,
every code path is fully unit-testable without mocking anything.
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl
import pytest

from app.schemas.inventory import ColumnMapping
from app.services import order_import_service as svc


# ---- guess_column_mapping -------------------------------------------------


def test_guess_column_mapping_matches_common_headers():
    headers = ["Order Date", "Item Name", "Qty", "Unit Price", "Category"]
    mapping = svc.guess_column_mapping(headers)
    assert mapping.name_column == "Item Name"
    assert mapping.quantity_column == "Qty"
    assert mapping.price_column == "Unit Price"
    assert mapping.date_column == "Order Date"
    # "Category" matches none of the keyword lists -- correctly left
    # unmapped rather than forced onto some field.
    assert mapping.unit_column is None


def test_guess_column_mapping_does_not_double_claim_a_header():
    # "Item Price" should be claimed by price_column, not name_column,
    # even though "item" is a name_column keyword -- name_column's own
    # keyword list is ordered to prefer an actual "item name"/"product"
    # style header over a bare substring match.
    headers = ["Item Price", "Product Description", "Quantity"]
    mapping = svc.guess_column_mapping(headers)
    assert mapping.name_column == "Product Description"
    assert mapping.price_column == "Item Price"


def test_guess_column_mapping_returns_all_none_for_unrecognized_headers():
    mapping = svc.guess_column_mapping(["Col A", "Col B"])
    assert mapping.model_dump() == {
        "name_column": None,
        "quantity_column": None,
        "unit_column": None,
        "price_column": None,
        "date_column": None,
    }


# ---- parse_tabular_file: CSV ----------------------------------------------


def test_parse_csv_basic():
    csv_bytes = b"Item Name,Qty,Price\nBananas,2,1.29\nMilk,1,3.49\n"
    headers, rows = svc.parse_tabular_file(csv_bytes, "order.csv", "text/csv")
    assert headers == ["Item Name", "Qty", "Price"]
    assert rows == [
        {"Item Name": "Bananas", "Qty": "2", "Price": "1.29"},
        {"Item Name": "Milk", "Qty": "1", "Price": "3.49"},
    ]


def test_parse_csv_strips_utf8_bom():
    csv_bytes = b"\xef\xbb\xbfItem Name,Qty\nBananas,2\n"
    headers, _ = svc.parse_tabular_file(csv_bytes, "order.csv", "text/csv")
    assert headers == ["Item Name", "Qty"]  # not "﻿Item Name"


def test_parse_unsupported_file_type_raises():
    with pytest.raises(ValueError, match="Unsupported file type"):
        svc.parse_tabular_file(b"whatever", "order.docx", "application/msword")


# ---- parse_tabular_file: XLSX ----------------------------------------------


def _build_xlsx(sheet_name: str, rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xlsx_basic():
    xlsx_bytes = _build_xlsx("Sheet1", [["Item Name", "Qty", "Price"], ["Bananas", 2, 1.29], ["Milk", 1, 3.49]])
    headers, rows = svc.parse_tabular_file(
        xlsx_bytes, "order.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert headers == ["Item Name", "Qty", "Price"]
    assert rows[0]["Item Name"] == "Bananas"
    assert rows[0]["Qty"] == "2"  # numeric cells come back as strings, same as the CSV path
    assert rows[1]["Item Name"] == "Milk"


def test_parse_xlsx_prefers_items_sheet_when_present():
    # Mirrors the Walmart Invoice Exporter extension's modern two-sheet
    # Orders+Items export layout (verified via its README during
    # research) -- "Items" is the one row-per-line-item sheet this
    # importer actually wants.
    wb = openpyxl.Workbook()
    orders_ws = wb.active
    orders_ws.title = "Orders"
    orders_ws.append(["Order Number", "Order Total"])
    orders_ws.append(["12345", "10.00"])
    items_ws = wb.create_sheet("Items")
    items_ws.append(["Item Name", "Qty"])
    items_ws.append(["Bananas", 2])
    buf = io.BytesIO()
    wb.save(buf)

    headers, rows = svc.parse_tabular_file(
        buf.getvalue(), "order.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert headers == ["Item Name", "Qty"]
    assert rows[0]["Item Name"] == "Bananas"


def test_parse_xlsx_skips_fully_blank_rows():
    xlsx_bytes = _build_xlsx("Sheet1", [["Item Name", "Qty"], ["Bananas", 2], [None, None], ["Milk", 1]])
    _, rows = svc.parse_tabular_file(
        xlsx_bytes, "order.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(rows) == 2


# ---- apply_mapping ----------------------------------------------------------


def test_apply_mapping_happy_path():
    headers = ["Item Name", "Qty", "Unit", "Price", "Order Date"]
    rows = [{"Item Name": "Bananas", "Qty": "2", "Unit": "lb", "Price": "$1.29", "Order Date": "2026-07-28"}]
    mapping = ColumnMapping(
        name_column="Item Name", quantity_column="Qty", unit_column="Unit", price_column="Price", date_column="Order Date"
    )
    items, skipped = svc.apply_mapping(headers, rows, mapping)
    assert skipped == 0
    assert len(items) == 1
    item = items[0]
    assert item.name == "Bananas"
    assert item.estimated_quantity == 2.0
    assert item.unit == "lb"
    assert item.unit_price == 1.29  # "$1.29" -> 1.29, currency symbol stripped
    assert item.purchased_date == date(2026, 7, 28)
    assert item.confidence_note is None


def test_apply_mapping_skips_rows_with_no_name():
    headers = ["Item Name", "Qty"]
    rows = [{"Item Name": "", "Qty": "1"}, {"Item Name": "Bananas", "Qty": "2"}]
    mapping = ColumnMapping(name_column="Item Name", quantity_column="Qty")
    items, skipped = svc.apply_mapping(headers, rows, mapping)
    assert skipped == 1
    assert len(items) == 1
    assert items[0].name == "Bananas"


def test_apply_mapping_defaults_quantity_to_one_when_unmapped():
    rows = [{"Item Name": "Bananas"}]
    mapping = ColumnMapping(name_column="Item Name")  # no quantity_column at all
    items, _ = svc.apply_mapping(["Item Name"], rows, mapping)
    assert items[0].estimated_quantity == 1.0
    assert "no quantity column mapped" in items[0].confidence_note


def test_apply_mapping_defaults_quantity_to_one_on_unparseable_value_and_notes_it():
    rows = [{"Item Name": "Bananas", "Qty": "several"}]
    mapping = ColumnMapping(name_column="Item Name", quantity_column="Qty")
    items, _ = svc.apply_mapping(["Item Name", "Qty"], rows, mapping)
    assert items[0].estimated_quantity == 1.0
    assert "could not parse quantity value 'several'" in items[0].confidence_note


def test_apply_mapping_leaves_price_none_on_unparseable_value_and_notes_it():
    rows = [{"Item Name": "Bananas", "Price": "N/A"}]
    mapping = ColumnMapping(name_column="Item Name", price_column="Price")
    items, _ = svc.apply_mapping(["Item Name", "Price"], rows, mapping)
    assert items[0].unit_price is None
    assert "could not parse price value 'N/A'" in items[0].confidence_note


def test_apply_mapping_strips_thousands_separator_from_price():
    rows = [{"Item Name": "Big order", "Price": "$1,234.56"}]
    mapping = ColumnMapping(name_column="Item Name", price_column="Price")
    items, _ = svc.apply_mapping(["Item Name", "Price"], rows, mapping)
    assert items[0].unit_price == 1234.56


def test_apply_mapping_leaves_date_none_when_unparseable():
    rows = [{"Item Name": "Bananas", "Order Date": "not a date"}]
    mapping = ColumnMapping(name_column="Item Name", date_column="Order Date")
    items, _ = svc.apply_mapping(["Item Name", "Order Date"], rows, mapping)
    assert items[0].purchased_date is None


def test_apply_mapping_splits_compound_unit_column_text():
    # Package/measurement split (2026-08-02): a spreadsheet's own "unit"
    # column is exactly as likely to say "8 oz" or "500g" as a receipt
    # line is. Qty=2 (how many purchased) * the parsed 8 oz package size
    # = 16 oz actually on hand.
    headers = ["Item Name", "Qty", "Unit"]
    rows = [{"Item Name": "Cheese", "Qty": "2", "Unit": "8 oz"}]
    mapping = ColumnMapping(name_column="Item Name", quantity_column="Qty", unit_column="Unit")
    items, _ = svc.apply_mapping(headers, rows, mapping)
    item = items[0]
    assert item.unit == "oz"
    assert item.package_quantity == 8
    assert item.package_count == 2
    assert item.estimated_quantity == 16


def test_apply_mapping_leaves_plain_unit_column_text_unsplit():
    # A clean unit with no leading number (e.g. "lb", already canonical)
    # has nothing to split -- package_quantity stays unset and
    # estimated_quantity is exactly the mapped quantity column value,
    # same as this importer's pre-existing behavior.
    headers = ["Item Name", "Qty", "Unit"]
    rows = [{"Item Name": "Rice", "Qty": "3", "Unit": "lb"}]
    mapping = ColumnMapping(name_column="Item Name", quantity_column="Qty", unit_column="Unit")
    items, _ = svc.apply_mapping(headers, rows, mapping)
    item = items[0]
    assert item.unit == "lb"
    assert item.package_quantity is None
    assert item.estimated_quantity == 3.0


def test_apply_mapping_parses_alternate_date_formats():
    mapping = ColumnMapping(name_column="Item Name", date_column="Order Date")
    for raw, expected in [
        ("07/28/2026", date(2026, 7, 28)),
        ("07/28/26", date(2026, 7, 28)),
        ("Jul 28, 2026", date(2026, 7, 28)),
    ]:
        items, _ = svc.apply_mapping(
            ["Item Name", "Order Date"], [{"Item Name": "Bananas", "Order Date": raw}], mapping
        )
        assert items[0].purchased_date == expected, f"failed for {raw!r}"
